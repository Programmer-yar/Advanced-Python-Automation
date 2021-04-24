import openpyxl

#add 'data_only=True' arg to get only value instead of formula
wb = openpyxl.load_workbook('calculator.xlsx', data_only=True)
ws = wb['Sheet1']

ws['A1'] = 120
ws['B1'] = 100

wb.save('calculator.xlsx')
