import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet_1 = wb['Sheet1']

#getting cell "A1" from sheet_1
a1 = sheet_1['A1']
print('Object of cell A1: ', a1)
print('value of cell A1: ', a1.value)

#getting cell "B1"
b1 = sheet_1['B1']
print('value of cell B1: ', b1.value)

#getting cell row
print('Row of cell A1 is: ', a1.row)

#coordinates of cell "columnRow"
print('coordinates of cell b1 are: ', b1.coordinate)

#'cell()' method to access row and columns
#'cell()' method accepts "row" and "column" as kwargs
c1 = sheet_1.cell(row=1, column=3)
print('value of cell c1: ', c1.value, '\n')


#traversing through cells and printing values
for i in range(1,4):
    print(sheet_1.cell(i, 2).value, sheet_1.cell(i,3).value)

print('-'*20)














